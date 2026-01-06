import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

OUTPUT_PATH = "data/results/v2/final_card_v2.xlsx"

def build_excel_card(df):

    # Fix team reversal
    # Player team should NOT equal opponent team
    # If they match, swap them.
    df["correct_team"] = df["player_team_name"]
    df["correct_opp"] = df["opp_team_name"]

    mask = df["player_team_name"] == df["opp_team_name"]
    df.loc[mask, "correct_team"] = df.loc[mask, "opp_team_name"]
    df.loc[mask, "correct_opp"] = df.loc[mask, "player_team_name"]

    df["player_team_name"] = df["correct_team"]
    df["opp_team_name"] = df["correct_opp"]

    df = df.drop(columns=["correct_team","correct_opp"], errors="ignore")

    # Split overs / unders
    overs = df[df["side"] == "over"].reset_index(drop=True)
    unders = df[df["side"] == "under"].reset_index(drop=True)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "NBA Prop Card"

    # -------------------------
    # HEADER
    # -------------------------
    date_str = datetime.now().strftime("%m/%d/%Y")
    header_text = f"@Jayssportsanalytics – NBA Player Prop Model – {date_str}"

    ws.merge_cells("A1:F1")
    header_cell = ws["A1"]
    header_cell.value = header_text
    header_cell.font = Font(size=16, bold=True, color="FFFFFF")  
    header_cell.fill = PatternFill("solid", fgColor="404040")  # DARK GREY
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # -------------------------
    # Styles
    # -------------------------
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    green_header = PatternFill("solid", fgColor="C6EFCE")
    red_header = PatternFill("solid", fgColor="FFC7CE")

    green_fill = PatternFill("solid", fgColor="E2F0D9")
    red_fill = PatternFill("solid", fgColor="F8CBAD")

    center = Alignment(horizontal="center", vertical="center")

    headers = ["Player", "Team", "Opponent", "Prop", "Odds", "AI Rating"]

    row_cursor = 3

    # -------------------------
    # OVERS TITLE
    # -------------------------
    ws.merge_cells(f"A{row_cursor}:F{row_cursor}")
    cell = ws[f"A{row_cursor}"]
    cell.value = "OVERS"
    cell.font = Font(size=14, bold=True)
    cell.fill = green_header
    cell.alignment = center
    row_cursor += 1

    # Column headers
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row_cursor, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = green_header
        c.alignment = center
        c.border = border

    row_cursor += 1

    # OVERS DATA
    for _, row in overs.iterrows():
        prop_name = row["market"].replace("_", " ").title()
        prop_text = f"Over {row['line']} {prop_name}"

        row_data = [
            row["player_name"],
            row["player_team_name"],
            row["opp_team_name"],
            prop_text,
            row["odds"],
            f"{row['confidence']*100:.1f}"
        ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_cursor, column=col, value=val)
            c.alignment = center
            c.fill = green_fill
            c.border = border

        row_cursor += 1

    # -------------------------
    # UNDERS TITLE
    # -------------------------
    ws.merge_cells(f"A{row_cursor}:F{row_cursor}")
    cell = ws[f"A{row_cursor}"]
    cell.value = "UNDERS"
    cell.font = Font(size=14, bold=True)
    cell.fill = red_header
    cell.alignment = center

    row_cursor += 1

    # Column headers
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row_cursor, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = red_header
        c.alignment = center
        c.border = border

    row_cursor += 1

    # UNDERS DATA
    for _, row in unders.iterrows():
        prop_name = row["market"].replace("_", " ").title()
        prop_text = f"Under {row['line']} {prop_name}"

        row_data = [
            row["player_name"],
            row["player_team_name"],
            row["opp_team_name"],
            prop_text,
            row["odds"],
            f"{row['confidence']*100:.1f}"
        ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_cursor, column=col, value=val)
            c.alignment = center
            c.fill = red_fill
            c.border = border

        row_cursor += 1

    # -------------------------
    # AUTO WIDTHS
    # -------------------------
    ws.column_dimensions["A"].width = 22  # Player
    ws.column_dimensions["B"].width = 20  # Team
    ws.column_dimensions["C"].width = 20  # Opponent
    ws.column_dimensions["D"].width = 32  # Prop
    ws.column_dimensions["E"].width = 12  # Odds
    ws.column_dimensions["F"].width = 12  # Rating

    os.makedirs("data/results/v2", exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[EXCEL UPDATED V2] Saved → {OUTPUT_PATH}")


if __name__ == "__main__":
    # Load the FINAL CARD (not raw sims)
    df = pd.read_csv("data/processed/final_card_today_v2.csv")
    print(f"[V2] Loaded {len(df)} picks for today.")

    build_excel_card(df)

